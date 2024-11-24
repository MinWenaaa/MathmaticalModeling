from PIL import Image
import os
import cv2
import numpy as np
from math import sqrt, log, isinf
from question1 import Bayes
from math import sqrt
import sys
from openpyxl import load_workbook

class IMGaccess:
    def __init__(self, file_path):
        image = Image.open(file_path)
        self.image_array = np.array(image)
        self.lab_image = cv2.cvtColor(self.image_array, cv2.COLOR_RGB2LAB).astype(np.float64)/255
        self.image_array = self.image_array.astype(np.float64)

    def PNSR(self, image_origin):
        MSE = np.mean((self.image_array-image_origin)**2)
        self.PNSR = 10 * log(255**2/MSE, 10)
        return self.PNSR

    def _UICM(self):
        R = self.image_array[:,:,0]
        G = self.image_array[:,:,1]
        B = self.image_array[:,:,2]
        # print('色彩均值：', np.mean(R), ", ", np.mean(G), ", ", np.mean(B))
        RG = (R - G).flatten()
        YB = ((R+G)/2-B).flatten()
        T_a = int(0.1*len(RG))
        mu_RG = np.mean(np.sort(RG)[T_a:-T_a])
        mu_YB = np.mean(np.sort(YB)[T_a:-T_a])
        omega_RG = np.sum((RG-mu_RG)**2)/len(RG)
        omega_YB = np.sum((YB-mu_YB)**2)/len(RG)
        self.UICM = -0.0268*np.sqrt(mu_RG**2+mu_YB**2) + 0.1586*np.sqrt(omega_RG+omega_YB)
        print("色彩测量指标(UICM): mu_RG: ", mu_RG, "mu_YB: ", mu_YB, "omege_RG: ", omega_RG, "omega_YB: ", omega_YB)
        return self.UICM

    def _UISM(self):
        h, w = self.image_array[:,:,0].shape
        self.UISM = 0
        for i in range(3):
            edge = self.image_array[:,:,i] #np.zeros((h-2,w-2))
            EME = 0
            for y in range(int((h-2)/40)):
                for x in range(int((w-2)/40)):
                    window = edge[y*40:y*40+40, x*40:x*40+40]
                    if (window.max()+0.1)/(window.min()+0.1) <= 0 or isinf((window.max()+0.1)/(window.min()+0.1)):
                        print(window, '\n', window.max(), ', ', window.min())
                        print((window.max()+0.1)/(window.min()+0.1))
                        sys.exit()
                    EME += log((window.max()+0.1)/(window.min()+1))
            self.UISM += EME
        self.UISM = self.UISM * 2 / (h/40) / (w/40)
        return self.UISM
    
    def _UIConM(self):
        bright_channel = np.amax(self.image_array, axis=2)
        h, w = self.image_array[:,:,0].shape
        AMEE = 0.0
        for y in range(int(h/40)):
            for x in range(int(w/40)):
                window = bright_channel[y*40:y*40+40, x*40:x*40+40]
                temp = (window.max()-window.min()+1)/(window.max()+window.min())
                AMEE += temp*log(temp) / (h/40) / (w/40)
                if isinf(AMEE):
                    print(window, '\n', window.max(), window.min())
                    print(temp, '\n', temp*log(temp))
                    sys.exit()
        self.UIConM = AMEE
        return self.UIConM
    
    def UIQM(self):
        self.uiqm = 0.0282 * self._UICM() + 0.4953 * self._UISM() + 2.5753 * self._UIConM()
        print("UICM: ", self.UICM, ", UISM: ", self.UISM, ", UIConM", self.UIConM,)
        return self.uiqm
    
    def _omega_c(self):
        C_i = np.sqrt(self.lab_image[:,:,1]**2 + self.lab_image[:,:,2]**2)
        C_bar = np.mean(C_i)
        h, w = C_i.shape
        self.omega_c = sqrt(np.sum((C_i-C_bar)**2)/h/w)
        return self.omega_c
    
    def _C_lum(self):
        L = self.lab_image[:,:,0].flatten()
        L = np.sort(L)
        index = int(0.01*len(L))
        self.C_lum = abs(L[index]-L[-index])
        return self.C_lum
    
    def _avg_sat(self):
        C_i = np.sqrt(self.lab_image[:,:,1]**2 + self.lab_image[:,:,2]**2)
        h,w = C_i.shape
        self.avg_sat = np.sum(C_i/(self.lab_image[:,:,0]+0.1))/(h*w)
        return self.avg_sat

    def UCIQE(self):
        self.uciqe = 0.4680*self._omega_c() + 0.2745*self._C_lum() + 0.2576*self._avg_sat()
        print("UCIQE: ", self.omega_c, ", ", self.C_lum, ", ", self.avg_sat, '\n')
        return self.uciqe

class image_correction:
    def __init__(self, file_path):
        image = Image.open(file_path)
        self.image_array = np.array(image)
        self.lab_image = cv2.cvtColor(self.image_array, cv2.COLOR_RGB2LAB).astype(np.float64)
        self.image_array = self.image_array.astype(np.float64)

    def color_correction(self):
        L = self.lab_image[:,:,0]
        A = self.lab_image[:,:,1]
        B = self.lab_image[:,:,2]
        d_L = L-(L.max()+L.min())/2
        L = (L.max()+L.min())/2 + 3*np.sign(d_L)*(abs(d_L)** (1/1.1))
        std_dev_A = 3* np.std(A)
        std_dev_B = 3* np.std(B)
        print(255/std_dev_A, 255/std_dev_B)
        A = (A-np.mean(A)) * ((255/std_dev_A)**0.4+0.1) + 128
        B = (B-np.mean(B)) * ((255/std_dev_B)**0.4+0.1) + 128

        new_lab = np.clip(np.stack((L,A,B), axis=2), 0, 255).astype(np.uint8)
        corrected = cv2.cvtColor(new_lab, cv2.COLOR_LAB2RGB)
        return corrected
    
    def bright_correct(self):
        L = self.lab_image[:,:,0]
        A = self.lab_image[:,:,1]
        B = self.lab_image[:,:,2]
        L = L + 0.5 * (128 - np.mean(L)) * np.sqrt(L/np.mean(L))
        A = np.mean(A) + 1.5 * (A-np.mean(A))
        B = np.mean(B) + 1.5 * (B-np.mean(B))
        new_lab = np.clip(np.stack((L,A,B), axis=2), 0, 255).astype(np.uint8)
        corrected = cv2.cvtColor(new_lab, cv2.COLOR_LAB2RGB)
        return corrected
    
    def image_shapenning(self):
        L = self.lab_image[:,:,0]
        A = self.lab_image[:,:,1]
        B = self.lab_image[:,:,2]
        L = np.mean(L) + 1.5*(L-np.mean(L))
        A = np.mean(A) + 1.5*(A-np.mean(A))
        B = np.mean(B) + 1.5*(B-np.mean(B))
        new_lab = np.clip(np.stack((L,A,B), axis=2), 0, 255).astype(np.uint8)
        new_rgb = cv2.cvtColor(new_lab, cv2.COLOR_LAB2RGB)
        h, w, channel = new_rgb.shape
        sharpenned = np.zeros((h,w,channel)).astype(np.float64)
        laplacian = np.array([[-0.5,-1,-0.5],[-1,6,-1],[-0.5,-1,-0.5]])
        for y in range(1, h-1):
            for x in range(1, w-1):
                d = new_rgb[y-1:y+2, x-1:x+2] * laplacian
                sharpenned[y,x] = new_rgb[y,x] + 0.8*np.sum(np.sum(d, axis=1), axis = 1)
                
        return np.clip(sharpenned, 0, 255).astype(np.uint8)

    def compositive_adjustment(self, classification:list):
        classification = np.array(classification)
        classification = (1/classification**3) / sum(1/classification**3)

        print("分类概率：", classification)

        L = self.lab_image[:,:,0]
        A = self.lab_image[:,:,1]
        B = self.lab_image[:,:,2]

        L = L + 0.5 * (138 - np.mean(L)) * np.sqrt(L/np.mean(L))
        A = A + (128 - np.mean(A))*0.8
        B = B + (128 - np.mean(B))*0.8

        std_dev_A = 3* np.std(A)
        std_dev_B = 3* np.std(B)
        A = ((A-np.mean(A)) * ((255/std_dev_A)**0.4+0.1) + 128) * classification[1]**0.5 + (1-classification[1]**0.5) * A
        B = ((B-np.mean(B)) * ((255/std_dev_B)**0.4+0.1) + 128) * classification[1]**0.5 + (1-classification[1]**0.5) * B
        d_L = L-(L.max()+L.min())/2
        L = ((L.max()+L.min())/2 + 3*np.sign(d_L)*(abs(d_L)** (1/1.1))) * (classification[1]+classification[0])**0.5 + (1-(classification[1]+classification[0])**0.5) * L

        new_lab = np.clip(np.stack((L,A,B), axis=2), 0, 255).astype(np.uint8)
        corrected = cv2.cvtColor(new_lab, cv2.COLOR_LAB2RGB).astype(np.float64)

        h, w, channel = new_lab.shape
        sharpenned = np.zeros((h,w,channel)).astype(np.float64)
        laplacian = np.array([[-0.5,-1,-0.5],[-1,6,-1],[-0.5,-1,-0.5]])
        for y in range(1, h-1):
            for x in range(1, w-1):
                d = corrected[y-1:y+2, x-1:x+2] * laplacian
                sharpenned[y,x] = corrected[y,x] + 0.8*np.sum(np.sum(d, axis=1), axis = 1)


        return np.clip(corrected, 0, 255).astype(np.uint8)

if __name__=="__main__":
    bayes = Bayes()
    bayes.read_model('bayes.model')
    work_root = 'Attachment 2'
    filename_list = os.listdir(work_root)
    workbook = load_workbook(filename="Answer.xlsx")
    worksheet = workbook["attachment1 results"]
    output_root = "question3_output"

    for index, file in enumerate(filename_list):
        print(file, "start correction")
        path = os.path.join(work_root, file)
        IMGcorrection = image_correction(path)
        image = Image.open(path)
        image_array = np.array(image)
        p = bayes.classification(image_array)

        match p.index(max(p)):
            case 0:
                corrected = IMGcorrection.image_shapenning().astype(np.uint8)
            case 1:
                corrected = IMGcorrection.color_correction().astype(np.uint8)
            case 2:
                corrected = IMGcorrection.bright_correct().astype(np.uint8)
        output_path = os.path.join(output_root, file)
        image = Image.fromarray(corrected, 'RGB')
        image.save(output_path)

        access = IMGaccess(output_path)
        worksheet.cell(row=index+2, column=1, value=file)
        worksheet.cell(row=index+2, column=2, value=bayes.num2label[p.index(max(p))])
        worksheet.cell(row=index+2, column=3, value=access.PNSR(image_array))
        worksheet.cell(row=index+2, column=4, value=access.UIQM())
        worksheet.cell(row=index+2, column=5, value=access.UCIQE())
    
    workbook.save('Answer.xlsx')