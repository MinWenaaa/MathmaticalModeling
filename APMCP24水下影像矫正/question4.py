from openpyxl import load_workbook
from PIL import Image
from question3 import image_correction, IMGaccess
from question1 import Bayes
import numpy as np
import os

if __name__=="__main__":
    bayes = Bayes()
    bayes.read_model('bayes.model')
    work_root = 'Attachment 2'
    filename_list = os.listdir(work_root)
    workbook = load_workbook(filename="Answer.xlsx")
    worksheet = workbook["attachment 2 results"]
    output_root = "question4_output"

    for index, file in enumerate(filename_list):
        print(file, "start correction")
        path = os.path.join(work_root, file)
        IMGcorrection = image_correction(path)
        image = Image.open(path)
        image_array = np.array(image)
        p = bayes.classification(image_array)
        corrected = IMGcorrection.compositive_adjustment(p)
        output_path = os.path.join(output_root, file)
        image = Image.fromarray(corrected, 'RGB')
        image.save(output_path)

        access = IMGaccess(output_path)
        worksheet.cell(row=index+2, column=1, value=file)
        worksheet.cell(row=index+2, column=2, value=access.PNSR(image_array))
        worksheet.cell(row=index+2, column=3, value=access.UIQM())
        worksheet.cell(row=index+2, column=4, value=access.UCIQE())
    
    workbook.save('Answer.xlsx')