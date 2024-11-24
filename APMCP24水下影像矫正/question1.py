from PIL import Image
from openpyxl import load_workbook
import numpy as np
import random
import math
import os

class Bayes:
    def __init__(self):
        self.label2num = {
            'blur': 0,
            'color cast': 1,
            'low light':2
        }
        self.num2label = ['blur', 'color cast', 'low light']

    def getfeature(self, image_array):
        laplacian = np.array([[0,1,0],[1,-4,1],[0,1,0]])
        data = np.zeros(4)
        point_num = 3600

        r = image_array[:, :, 0].flatten()
        g = image_array[:, :, 1].flatten()
        b = image_array[:, :, 2].flatten()
        means = [np.mean(r), np.mean(g), np.mean(b)]
        data[0] = max(means)
        data[1] = max(abs(means[0] - means[1]), abs(means[1] - means[2]), abs(means[2] - means[0]))
        data[2] = np.mean([np.var(r), np.var(g), np.var(b)])

        h = image_array.shape[0]-3
        w = image_array.shape[1]-3
        random_numbers = random.sample(range(h*w), point_num)
        x = [int(i/w) for i in random_numbers] 
        y = [int(i%w) for i in random_numbers]
        edge_list = [None] * point_num
        for i, (temp_x, temp_y) in enumerate(zip(x, y)):
            sample1 = image_array[temp_x:temp_x+3, temp_y:temp_y+3, 0]
            sample2 = image_array[temp_x:temp_x+3, temp_y:temp_y+3, 1]
            sample3 = image_array[temp_x:temp_x+3, temp_y:temp_y+3, 2]

            edge_list[i] = abs(np.sum(sample1 * laplacian))+abs(np.sum(sample2 * laplacian))+abs(np.sum(sample3 * laplacian))
        edge_sorted = np.sort(edge_list)
        edge_sorted = edge_sorted[-250:]
        data[3] = np.mean(edge_sorted)
        return data

    def train(self, train_list:list[str]=[]):
        train_set = np.zeros((len(train_list), 4))

        for index, filename in enumerate(train_list):
            image = Image.open(filename)
            image_array = np.array(image)
            train_set[index] = self.getfeature(image_array)

        num = [0,0,0]
        for key in train_list:
            num[self.label2num[key.split("\\")[1]]] += 1

        self.means1 = np.mean(train_set[:num[0],:], axis=0)
        self.means2 = np.mean(train_set[num[0]:num[0]+num[1],:], axis=0)
        self.means3 = np.mean(train_set[num[0]+num[1]:,:], axis=0)

        cov1 = np.cov(train_set[:num[0],:], rowvar=False)
        cov2 = np.cov(train_set[num[0]:num[0]+num[1],:], rowvar=False)
        cov3 = np.cov(train_set[num[0]+num[1]:,:], rowvar=False)

        self.Inv1 = np.linalg.inv(cov1)# 计算协方差的逆
        self.Inv2 = np.linalg.inv(cov2)
        self.Inv3 = np.linalg.inv(cov3)
        self.Det1 = np.linalg.det(cov1)# 计算协方差的行列式
        self.Det2 = np.linalg.det(cov2)
        self.Det3 = np.linalg.det(cov3)

    def classification(self, image_array):
        feature = self.getfeature(image_array)

        A = np.vstack((feature-self.means1, feature-self.means2, feature-self.means3))
        M1 = np.dot(np.dot(A[0],self.Inv1),A.T[:,0])
        M2 = np.dot(np.dot(A[1],self.Inv2),A.T[:,1])
        M3 = np.dot(np.dot(A[2],self.Inv3),A.T[:,2])

        P = [-0.5 * ((math.log(self.Det1)) + M1 + 4*math.log(2 * math.pi)),
                -0.5 * ((math.log(self.Det2)) + M2 + 4*math.log(2 * math.pi)),
                -0.5 * ((math.log(self.Det3)) + M3 + 4*math.log(2 * math.pi))]

        return P
    
    def save_model(self, output_file):
        model_data = np.vstack((
            self.means1, self.means2, self.means3,
            self.Inv1, self.Inv2, self.Inv3,
            [self.Det1, self.Det2, self.Det3, 0]
        )).astype(np.float64)

        np.savetxt(output_file, model_data, delimiter=',', fmt='%f')

    def read_model(self, file_path:str):
        data = np.loadtxt(file_path, delimiter=',')
        self.means1 = data[0,:]
        self.means2 = data[1,:]
        self.means3 = data[2,:]
        self.Inv1 = data[3:7,:]
        self.Inv2 = data[7:11,:]
        self.Inv3 = data[11:15,:]
        self.Det1 = data[15,0]
        self.Det2 = data[15,1]
        self.Det3 = data[15,2]

        

if __name__=="__main__":

    # train_list = [
    #     os.path.join('train', 'blur', filename) for filename in os.listdir('train\\blur')
    # ] + [
    #     os.path.join('train', 'color_cast', filename) for filename in os.listdir('train\\color_cast')
    # ] + [
    #     os.path.join('train', 'low_light', filename) for filename in os.listdir('train\\low_light')
    # ]

    # bayes = Bayes()
    # bayes.train(train_list)
    # bayes.save_model('bayes.model')


    bayes = Bayes()
    bayes.read_model('bayes.model')

    workbook = load_workbook(filename="Answer.xlsx")
    worksheet = workbook["question1"]
    work_root = 'Attachment 1'
    filename_list = os.listdir(work_root)
    for index, file in enumerate(filename_list):
        path = os.path.join(work_root, file)
        image = Image.open(path)
        image_array = np.array(image)
        p = bayes.classification(image_array)
        result = bayes.num2label[p.index(max(p))]
        worksheet.cell(row=index+2, column=1, value=file)
        worksheet.cell(row=index+2, column=2, value=result)

    workbook.save('Answer.xlsx')

